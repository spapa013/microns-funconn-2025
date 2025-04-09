from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import shutil


def prepare_rslt_dir(dir, remove_existing=False):
    dir = Path(dir)
    if not dir.exists():
        dir.mkdir(parents=True)
    # remove existing files and subdirectories
    if remove_existing:
        for f in dir.iterdir():
            if f.is_file():
                f.unlink()
            elif f.is_dir():
                shutil.rmtree(f)
    return dir


def to_excel(df: pd.DataFrame, path, docstring):
    """save dataframe to excel file, with docstring appended to the fisrt row

    Parameters
    ----------
    df : pd.DataFrame
        dataframe to save
    path : path-like
        path to save the file
    docstring : str
        docstring to append to the first row
    """
    df.to_excel(path, engine="openpyxl", startrow=2, index=False)
    wb = load_workbook(path)
    ws = wb.active
    ws["A1"] = docstring
    wb.save(path)
    return path


def concat_excel(files, path):
    """concatenate multiple excel files into one, each file is a sheet, sheet name is the file name

    Parameters
    ----------
    files : list
        list of path-like
    path : path-like
        path to save the concatenated file
    """
    with pd.ExcelWriter(path) as writer:
        for f in files:
            df = pd.read_excel(f)
            df.to_excel(writer, sheet_name=f.stem, index=False)
    return path
